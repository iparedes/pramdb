from openpyxl import load_workbook
from openpyxl.utils.cell import *
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


class AssessmentXLS:

    def __init__(self,filename):
        self.filename=filename
        self.book = load_workbook(filename)
        self.ws = self.book['Assessment']
        self.book.active=self.ws


        r=self.book.defined_names['AssetName']
        dests=list(r.destinations)[0][1]
        self.assetName=self.ws[dests].value

        r=self.book.defined_names['AssetType']
        dests=list(r.destinations)[0][1]
        self.assetType=self.ws[dests].value

        self.impacts = self.table_to_list('TblImpact')
        self.asls = self.table_to_list('TblAssessment')
        self.scenarios=self.table_to_list('TblScenarios')

    # Needs a table identifier
    # Returns a list of where each element corresponds to a row in the list, being a dictionary with the title of the columns as keys
    # if the first column is blank, skips the row
    def table_to_list(self, tableId,sheetname=None):

        if sheetname:
            ws=self.book.get_sheet_by_name(sheetname)
        else:
            ws=self.book.active

        tab = ws.tables[tableId]
        titles = tab.column_names
        # range=self.book.defined_names['TblImpact'].value
        rango = ws.tables[tableId].ref
        pattern = '([A-Z]+)([0-9]+)\:([A-Z]+)([0-9]+)'
        res = re.match(pattern, rango)

        a = res.group(1)
        colIni = column_index_from_string(a)

        rowIni = int(res.group(2))
        rowIni += 1  # skip the title

        a = res.group(3)
        colEnd = column_index_from_string(a)

        rowEnd = int(res.group(4))

        records = []
        for row in range(rowIni, rowEnd + 1):
            item = {}
            cont = 0
            for t in titles:
                item[t] = ws.cell(row, colIni + cont).value
                cont += 1
            if item[titles[0]]:
                records.append(item)
        return records

    def range_char(self, start, stop):
        return (chr(n) for n in range(ord(start), ord(stop) + 1))


    # Creates a sheet with a given name
    # Overwrites the sheet if it already exists
    # sets the new sheet to the active sheet
    def create_sheet(self,name):
        snames=self.book.sheetnames
        for s in snames:
            if s==name:
                sheet=self.book.get_sheet_by_name(name)
                self.book.remove_sheet(sheet)
        new_sheet=self.book.create_sheet(name)
        self.book.save(self.filename)

    def set_cell(self,row,column,value,sheetname=""):
        if sheetname:
            ws=self.book.get_sheet_by_name(sheetname)
        else:
            ws=self.book.active
        ws.cell(row,column).value=value

    def add_vector(self,row,column,vector,sheetname=""):
        if sheetname:
            ws=self.book.get_sheet_by_name(sheetname)
        else:
            ws=self.book.active
        col=column

        for v in vector:
            # r=type(v)
            # if r==str:
            #     t='s'
            # else:
            #     t='m'
            ws.cell(row=row,column=col).value=v
            # ws.cell(row=row, column=col).data_type=t
            #ws.cell(row=row, column=col).data_type = "s"
            # cell=ws.cell(row=row,column=col)
            col+=1
        #self.book.save("assessment.xlsx")


    def create_table(self,name,row,column,rows,sheetname=""):
        if not sheetname:
            sheetname=self.book.active.title

        rown=row
        for r in rows:
            self.add_vector(rown,column,r,sheetname)
            rown+=1

        StartColLetter=get_column_letter(column)
        EndColumn=column+len(rows[0])-1
        EndColLetter=get_column_letter(EndColumn)
        range= StartColLetter + str(row) + ":" + EndColLetter + str(rown-1)

        ws = self.book[sheetname]
        try:
            del ws.tables[name]
        except:
            pass
        tab=Table(displayName=name,name=name,ref=range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

    def delete_table(self,name):
        for s in self.book.worksheets:
            if name in s.tables.keys():
                t=s.tables[name]
                range=t.ref
                rows=rows_from_range(range)
                for r in rows:
                    for c in r:
                        coord=coordinate_from_string(c)
                        col=column_index_from_string(coord[0])
                        row=coord[1]
                        s.cell(row=row,column=col).value=None
                del s.tables[name]




    def add_data_validation(self,row,col,lista,sheetname=""):
        if sheetname:
            ws=self.book.get_sheet_by_name(sheetname)
        else:
            ws=self.book.active

        hop='"'+lista+'"'
        #dv = DataValidation(type='list', formula1='"poor, fair, good, no mets"', allowBlank=False, showDropDown=False)
        dv = DataValidation(type='list', formula1=hop, allowBlank=False, showDropDown=False)
        column=get_column_letter(col)
        cell=column+str(row)
        dv.add(cell)
        ws.add_data_validation(dv)
        #dv.ranges.ranges.append(cell)








